# src/report_builder.py
"""
Human-facing outputs for Data Analysis / Data Science - 3 deliverables:

1. build_dashboard()     -> results/dashboard.html   (business analytics)
2. build_model_report()  -> results/model_report.html (model diagnostics)
3. build_excel_report()  -> results/reports/forecast_report.xlsx

Design rules applied (so the pages stay clean if you edit them):
- Chart titles/subtitles live in the card's HTML, NOT inside Plotly - they
  wrap on narrow screens and free the strip above the plot for the legend
  (top-left) and filter dropdown (top-right), so nothing overlaps.
- One label rule: annotate selectively (a single subtitle note), never a
  text label on every mark.
- All half-width charts share one height; sections are labeled groups,
  not an unbroken wall of charts.

NOTE on units: Weekly_Sales in the Walmart dataset is measured in DOLLARS,
not item counts - all quantities are presented as $ value.
"""

import numpy as np
import pandas as pd
from pathlib import Path
from datetime import datetime
from .utils import logger
from .run_history import load_runs

try:
    import plotly.graph_objects as go
except ImportError:
    go = None

# ==========================================================================
# PALETTE (dataviz standard: fixed categorical order, single-hue sequential)
# ==========================================================================
INK = '#0b0b0b'
INK_2 = '#52514e'
MUTED = '#898781'
GRID = '#e1e0d9'
BASELINE = '#c3c2b7'
SURFACE = '#fcfcfb'
PLANE = '#f9f9f7'
SERIES_1 = '#2a78d6'   # blue  - Actual / primary
SERIES_2 = '#1baf7a'   # aqua  - Predicted / secondary
SEQ = ['#cde2fb', '#9ec5f4', '#6da7ec', '#3987e5', '#256abf', '#184f95', '#0d366b']
STATUS = {'good': '#0ca30c', 'warning': '#fab219', 'serious': '#ec835a', 'critical': '#d03b3b'}

FONT = 'system-ui, -apple-system, "Segoe UI", sans-serif'

H_HALF = 340   # uniform height for half-width charts
H_FULL = 380   # full-width charts


# ==========================================================================
# SHARED HELPERS
# ==========================================================================

def _base_layout(fig, height=H_HALF, top=28):
    """No in-plot title: titles live in the card HTML above the chart."""
    fig.update_layout(
        paper_bgcolor=SURFACE, plot_bgcolor=SURFACE,
        font=dict(family=FONT, size=12, color=INK_2),
        margin=dict(l=52, r=12, t=top, b=40),
        height=height,
        hovermode='x unified',
        showlegend=False,
        xaxis=dict(gridcolor=GRID, linecolor=BASELINE, zeroline=False,
                   tickfont=dict(color=MUTED, size=11)),
        yaxis=dict(gridcolor=GRID, linecolor=BASELINE, zeroline=False,
                   tickfont=dict(color=MUTED, size=11)),
    )
    return fig


def _fig_html(fig, include_js=False):
    return fig.to_html(
        full_html=False,
        include_plotlyjs='inline' if include_js else False,
        config={'displayModeBar': False, 'responsive': True}
    )


def _card(title, subtitle, fig, size='half', include_js=False, note=None):
    """Chart card: HTML title + subtitle (wrap on narrow screens), then plot."""
    sub = f'<p>{subtitle}</p>' if subtitle else ''
    note_html = f'<div class="note">{note}</div>' if note else ''
    return f'''<div class="card{' full' if size == 'full' else ''}">
      <div class="chead"><h2>{title}</h2>{sub}</div>
      {_fig_html(fig, include_js)}
      {note_html}
    </div>'''


def _kpi_card(label, value, sub='', tone=None):
    accent = f'border-left: 3px solid {STATUS[tone]};' if tone else ''
    sub_html = f'<div class="kpi-sub">{sub}</div>' if sub else ''
    return f'''<div class="kpi" style="{accent}">
      <div class="kpi-label">{label}</div>
      <div class="kpi-value">{value}</div>
      {sub_html}
    </div>'''


def _section(label, inner):
    return f'<div class="section-label">{label}</div>\n<div class="grid">{inner}</div>'


def _page_css():
    return f'''
  * {{ box-sizing: border-box; margin: 0; }}
  body {{ background: {PLANE}; font-family: {FONT}; color: {INK}; padding: 28px 24px; }}
  .wrap {{ max-width: 1280px; margin: 0 auto; }}
  header {{ display: flex; justify-content: space-between; align-items: flex-start;
            flex-wrap: wrap; gap: 12px; }}
  header .eyebrow {{ font-size: 12px; font-weight: 600; letter-spacing: 0.08em;
                     text-transform: uppercase; color: {SERIES_1}; }}
  header h1 {{ font-size: 24px; font-weight: 650; margin-top: 2px; }}
  header .meta {{ color: {MUTED}; font-size: 13px; margin-top: 6px; }}
  nav a {{ font-size: 13px; color: {SERIES_1}; text-decoration: none; font-weight: 600;
           border: 1px solid {GRID}; border-radius: 8px; padding: 8px 14px;
           background: {SURFACE}; display: inline-block; }}
  nav a:hover {{ border-color: {SERIES_1}; }}
  .section-label {{ font-size: 12px; font-weight: 600; letter-spacing: 0.08em;
                    text-transform: uppercase; color: {MUTED}; margin: 28px 0 10px; }}
  .kpis {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
           gap: 12px; }}
  .kpi {{ background: {SURFACE}; border: 1px solid rgba(11,11,11,0.10);
          border-radius: 12px; padding: 14px 16px; }}
  .kpi-label {{ font-size: 12px; color: {MUTED}; }}
  .kpi-value {{ font-size: 26px; font-weight: 650; margin-top: 4px; }}
  .kpi-sub {{ font-size: 12px; color: {INK_2}; margin-top: 2px; }}
  .insights {{ background: {SURFACE}; border: 1px solid rgba(11,11,11,0.10);
               border-left: 3px solid {SERIES_1}; border-radius: 12px;
               padding: 16px 20px; }}
  .insights h2 {{ font-size: 14px; font-weight: 650; margin-bottom: 8px; }}
  .insights li {{ font-size: 13.5px; color: {INK_2}; line-height: 1.65;
                  margin-left: 18px; margin-bottom: 4px; }}
  .insights b {{ color: {INK}; }}
  .grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 14px; }}
  .card {{ background: {SURFACE}; border: 1px solid rgba(11,11,11,0.10);
           border-radius: 12px; padding: 0 0 6px; overflow: hidden; }}
  .card.full {{ grid-column: 1 / -1; }}
  .card .chead {{ padding: 14px 16px 4px; }}
  .card .chead h2 {{ font-size: 15px; font-weight: 650; }}
  .card .chead p {{ font-size: 12px; color: {MUTED}; margin-top: 3px; line-height: 1.5; }}
  .card h2.plain {{ font-size: 15px; font-weight: 650; padding: 14px 16px 4px; }}
  .card .pad {{ padding: 8px 14px 14px; }}
  .tbl {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  .tbl th {{ text-align: left; color: {MUTED}; font-weight: 600; font-size: 12px;
             padding: 8px 10px; border-bottom: 1px solid {GRID}; white-space: nowrap; }}
  .tbl td {{ padding: 8px 10px; border-bottom: 1px solid {GRID}; }}
  .tbl .num {{ text-align: right; font-variant-numeric: tabular-nums; }}
  .tbl th.num {{ text-align: right; }}
  .chip {{ font-size: 12px; font-weight: 600; color: {INK};
           background: color-mix(in srgb, var(--c) 18%, white);
           border: 1px solid var(--c); border-radius: 99px; padding: 2px 10px;
           white-space: nowrap; }}
  .empty {{ color: {MUTED}; padding: 16px; }}
  .note {{ font-size: 12px; color: {MUTED}; padding: 4px 16px 8px; line-height: 1.5; }}
  footer {{ color: {MUTED}; font-size: 12px; margin: 26px 0 8px; line-height: 1.7; }}
  @media (max-width: 900px) {{ .grid {{ grid-template-columns: 1fr; }} }}
'''


def _fmt_money(v, decimals=0):
    return f'${v:,.{decimals}f}'


def _fmt_pct(v, decimals=1):
    return '—' if v is None or (isinstance(v, float) and np.isnan(v)) else f'{v:.{decimals}f}%'


def _eval_df(eval_data):
    return pd.DataFrame({
        'Date': pd.to_datetime(eval_data['Date']),
        'Store': eval_data['Store'],
        'Dept': eval_data['Dept'],
        'Cluster': eval_data.get('Cluster', np.zeros(len(eval_data['y_true']), dtype=int)),
        'Actual': np.asarray(eval_data['y_true'], dtype=float),
        'Predicted': np.asarray(eval_data['y_pred'], dtype=float),
        'IsHoliday': np.asarray(eval_data['is_holiday']).astype(bool)
                     if 'is_holiday' in eval_data else np.zeros(len(eval_data['y_true']), dtype=bool),
    })


def _add_holiday_lines(fig, df):
    """Dotted verticals only - the meaning is explained ONCE in the subtitle."""
    for d in sorted(df.loc[df['IsHoliday'], 'Date'].unique()):
        fig.add_vline(x=d, line_width=1, line_dash='dot', line_color=STATUS['warning'])
    return fig


def _bar(fig):
    """Common bar polish: labels never clipped."""
    fig.update_traces(cliponaxis=False)
    return fig


# ==========================================================================
# CHARTS - ANALYST PAGE
# ==========================================================================

def _chart_forecast_timeseries(df):
    fig = go.Figure()
    groups = [('All clusters', None)] + [(f'Cluster {c}', c) for c in sorted(pd.unique(df['Cluster']))]

    for label, cluster in groups:
        sub = df if cluster is None else df[df['Cluster'] == cluster]
        weekly = sub.groupby('Date')[['Actual', 'Predicted']].sum().reset_index()
        visible = cluster is None
        fig.add_trace(go.Scatter(
            x=weekly['Date'], y=weekly['Actual'], name='Actual',
            line=dict(color=SERIES_1, width=2), visible=visible,
            hovertemplate='Actual: $%{y:,.0f}<extra></extra>'))
        fig.add_trace(go.Scatter(
            x=weekly['Date'], y=weekly['Predicted'], name='Forecast',
            line=dict(color=SERIES_2, width=2, dash='dash'), visible=visible,
            hovertemplate='Forecast: $%{y:,.0f}<extra></extra>'))

    buttons = []
    for i, (label, _) in enumerate(groups):
        vis = [False] * (2 * len(groups))
        vis[2 * i] = vis[2 * i + 1] = True
        buttons.append(dict(label=label, method='update', args=[{'visible': vis}]))

    _base_layout(fig, height=H_FULL, top=48)
    _add_holiday_lines(fig, df)
    # Freed strip above the plot: legend on the LEFT, filter on the RIGHT
    fig.update_layout(
        showlegend=True,
        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='left', x=0,
                    font=dict(size=11, color=INK_2)),
        updatemenus=[dict(
            buttons=buttons, direction='down',
            x=1.0, xanchor='right', y=1.18, yanchor='top',
            bgcolor=SURFACE, bordercolor=GRID, font=dict(size=11, color=INK_2),
            pad=dict(t=0, b=0))],
        yaxis_tickformat='~s', yaxis_tickprefix='$',
    )
    return fig


def _chart_error_by_cluster(df):
    g = df.assign(err=(df['Actual'] - df['Predicted']).abs()).groupby('Cluster')['err'].mean().reset_index()
    fig = go.Figure(go.Bar(
        x=[f'Cluster {c}' for c in g['Cluster']], y=g['err'],
        marker=dict(color=SEQ[3], cornerradius=4),
        text=[f'{v:,.0f}' for v in g['err']], textposition='outside',
        textfont=dict(color=INK_2, size=11),
        hovertemplate='MAE: $%{y:,.0f}<extra></extra>', width=0.55,
    ))
    _base_layout(fig)
    fig.update_layout(hovermode='closest', yaxis_tickformat='~s', yaxis_tickprefix='$')
    return _bar(fig)


def _chart_store_scatter(df):
    per_store = df.groupby('Store')[['Actual', 'Predicted']].sum().reset_index()
    lim = max(per_store['Actual'].max(), per_store['Predicted'].max()) * 1.05

    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=[0, lim], y=[0, lim], mode='lines',
        line=dict(color=BASELINE, width=1, dash='dot'), hoverinfo='skip'))
    fig.add_trace(go.Scatter(
        x=per_store['Actual'], y=per_store['Predicted'], mode='markers',
        marker=dict(color=SERIES_1, size=9, opacity=0.85, line=dict(color=SURFACE, width=1)),
        customdata=per_store['Store'],
        hovertemplate='Store %{customdata}<br>Actual: $%{x:,.0f}<br>Forecast: $%{y:,.0f}<extra></extra>'))

    _base_layout(fig)
    fig.update_layout(
        hovermode='closest',
        xaxis=dict(title=dict(text='Actual ($)', font=dict(size=11, color=MUTED)), tickformat='~s'),
        yaxis=dict(title=dict(text='Forecast ($)', font=dict(size=11, color=MUTED)), tickformat='~s'),
    )
    return fig


def _chart_abc_xyz(store_dept_full):
    g = store_dept_full.groupby(['ABC', 'XYZ']).agg(
        n=('Store', 'size'), sales=('Total_Actual_Sales', 'sum')).reset_index()
    total_sales = g['sales'].sum()
    pivot_n = g.pivot(index='ABC', columns='XYZ', values='n').reindex(
        index=['A', 'B', 'C'], columns=['X', 'Y', 'Z']).fillna(0)
    pivot_s = g.pivot(index='ABC', columns='XYZ', values='sales').reindex(
        index=['A', 'B', 'C'], columns=['X', 'Y', 'Z']).fillna(0) / (total_sales + 1e-9) * 100

    text = [[f'{int(pivot_n.iloc[i, j])}<br><span style="font-size:10px">{pivot_s.iloc[i, j]:.0f}% rev</span>'
             for j in range(3)] for i in range(3)]

    fig = go.Figure(go.Heatmap(
        z=pivot_n.values,
        x=['X (stable)', 'Y (variable)', 'Z (erratic)'],
        y=['A (70% revenue)', 'B (20%)', 'C (10%)'],
        colorscale=[[0, SEQ[0]], [1, SEQ[6]]],
        text=text, texttemplate='%{text}', textfont=dict(size=13),
        customdata=np.round(pivot_s.values, 1),
        hovertemplate='Class %{y} × %{x}<br>%{z} Store-Dept pairs · %{customdata}% of revenue<extra></extra>',
        showscale=False, xgap=2, ygap=2,
    ))
    _base_layout(fig)
    fig.update_layout(hovermode='closest', yaxis=dict(autorange='reversed'))
    return fig


def _chart_error_concentration(df, top_n=10):
    g = (df.assign(abs_err=(df['Actual'] - df['Predicted']).abs())
           .groupby('Dept')['abs_err'].sum().sort_values(ascending=False))
    total = g.sum()
    top = g.head(top_n)[::-1]

    fig = go.Figure(go.Bar(
        x=top.values, y=[f'Dept {int(d)}' for d in top.index], orientation='h',
        marker=dict(color=SEQ[3], cornerradius=4),
        text=[f'{v / total * 100:.1f}%' for v in top.values], textposition='outside',
        textfont=dict(color=INK_2, size=11),
        hovertemplate='%{y}: $%{x:,.0f} total error<extra></extra>',
    ))
    _base_layout(fig)
    fig.update_layout(hovermode='closest',
                      xaxis_tickformat='~s', xaxis_tickprefix='$',
                      margin=dict(l=76, r=48, t=28, b=40))
    return _bar(fig)


# ==========================================================================
# CHARTS - MODEL DIAGNOSTICS PAGE
# ==========================================================================

def _chart_model_comparison(df_cmp):
    df = df_cmp.sort_values('WMAE')
    has_detail = {'MAE', 'MAPE'}.issubset(df.columns)
    fig = go.Figure(go.Bar(
        x=df['Model'], y=df['WMAE'],
        marker=dict(color=[SERIES_1 if i == 0 else SEQ[1] for i in range(len(df))], cornerradius=4),
        text=[f'{v:,.0f}' for v in df['WMAE']], textposition='outside',
        textfont=dict(color=INK_2, size=11),
        customdata=df[['MAE', 'MAPE']].values if has_detail else None,
        hovertemplate=('%{x}<br>WMAE: %{y:,.0f}<br>MAE: %{customdata[0]:,.0f} · '
                       'MAPE: %{customdata[1]:.1f}%<extra></extra>') if has_detail
                      else '%{x}: WMAE %{y:,.0f}<extra></extra>',
        width=0.6,
    ))
    _base_layout(fig)
    fig.update_layout(hovermode='closest')
    return _bar(fig)


def _chart_train_time(df_cmp):
    if 'Train_Time_s' not in df_cmp.columns:
        return None
    df = df_cmp.sort_values('Train_Time_s')
    fig = go.Figure(go.Bar(
        x=df['Model'], y=df['Train_Time_s'],
        marker=dict(color=SEQ[2], cornerradius=4),
        text=[f'{v:,.0f}s' for v in df['Train_Time_s']], textposition='outside',
        textfont=dict(color=INK_2, size=11),
        hovertemplate='%{x}: %{y:.1f}s<extra></extra>', width=0.6,
    ))
    _base_layout(fig)
    fig.update_layout(hovermode='closest')
    return _bar(fig)


def _chart_residual_hist(df, clip=5000):
    resid = df['Actual'] - df['Predicted']
    clipped = resid.clip(-clip, clip)

    fig = go.Figure(go.Histogram(
        x=clipped, nbinsx=80, marker=dict(color=SEQ[3]),
        hovertemplate='Residual %{x}<br>%{y:,} observations<extra></extra>',
    ))
    fig.add_vline(x=0, line_width=1, line_color=BASELINE)
    med = float(resid.median())
    fig.add_vline(x=med, line_width=1.5, line_dash='dash', line_color=STATUS['serious'])

    _base_layout(fig)
    fig.update_layout(hovermode='closest', bargap=0.05,
                      xaxis_tickformat='~s', xaxis_tickprefix='$')
    return fig


def _chart_mae_by_week(df):
    g = (df.assign(abs_err=(df['Actual'] - df['Predicted']).abs())
           .groupby('Date')['abs_err'].mean().reset_index())
    fig = go.Figure(go.Scatter(
        x=g['Date'], y=g['abs_err'], mode='lines+markers',
        line=dict(color=SERIES_1, width=2), marker=dict(size=6),
        hovertemplate='MAE: $%{y:,.0f}<extra></extra>'))
    _base_layout(fig, height=H_FULL)
    _add_holiday_lines(fig, df)
    fig.update_layout(yaxis_tickformat='~s', yaxis_tickprefix='$')
    return fig


def _chart_holiday_error(df):
    g = (df.assign(abs_err=(df['Actual'] - df['Predicted']).abs())
           .groupby('IsHoliday')['abs_err'].agg(['mean', 'size']).reset_index())
    labels = {False: 'Regular weeks', True: 'Holiday weeks'}
    x = [labels[b] for b in g['IsHoliday']]
    fig = go.Figure(go.Bar(
        x=x, y=g['mean'],
        marker=dict(color=[SEQ[2] if not b else STATUS['warning'] for b in g['IsHoliday']],
                    cornerradius=4),
        text=[f'{v:,.0f}' for v in g['mean']], textposition='outside',
        textfont=dict(color=INK_2, size=11),
        customdata=g['size'], width=0.45,
        hovertemplate='%{x}<br>MAE: $%{y:,.0f} · %{customdata:,} observations<extra></extra>',
    ))
    _base_layout(fig)
    fig.update_layout(hovermode='closest', yaxis_tickformat='~s', yaxis_tickprefix='$')
    return _bar(fig)


def _chart_residual_vs_predicted(df, sample=8000):
    d = df if len(df) <= sample else df.sample(sample, random_state=42)
    resid = d['Actual'] - d['Predicted']
    fig = go.Figure(go.Scattergl(
        x=d['Predicted'], y=resid, mode='markers',
        marker=dict(color=SERIES_1, size=4, opacity=0.25),
        hovertemplate='Forecast: $%{x:,.0f}<br>Residual: $%{y:,.0f}<extra></extra>',
    ))
    fig.add_hline(y=0, line_width=1, line_color=BASELINE)
    _base_layout(fig)
    fig.update_layout(hovermode='closest',
                      xaxis=dict(title=dict(text='Forecast value ($)', font=dict(size=11, color=MUTED)),
                                 tickformat='~s'),
                      yaxis=dict(title=dict(text='Residual ($)', font=dict(size=11, color=MUTED)),
                                 tickformat='~s'))
    return fig


def _chart_worst_depts(df, top_n=10):
    g = (df.assign(abs_err=(df['Actual'] - df['Predicted']).abs())
           .groupby('Dept').agg(mae=('abs_err', 'mean'), n=('abs_err', 'size')))
    g = g[g['n'] >= 30].sort_values('mae', ascending=False).head(top_n)[::-1]
    fig = go.Figure(go.Bar(
        x=g['mae'], y=[f'Dept {int(d)}' for d in g.index], orientation='h',
        marker=dict(color=SEQ[3], cornerradius=4),
        customdata=g['n'],
        hovertemplate='%{y}<br>MAE: $%{x:,.0f} · %{customdata:,} observations<extra></extra>',
    ))
    _base_layout(fig)
    fig.update_layout(hovermode='closest',
                      xaxis_tickformat='~s', xaxis_tickprefix='$',
                      margin=dict(l=76, r=24, t=28, b=40))
    return _bar(fig)


def _chart_feature_importance(df_imp, model_name, top_n=15):
    top = df_imp.head(top_n).sort_values('Importance')
    fig = go.Figure(go.Bar(
        x=top['Importance'], y=top['Feature'], orientation='h',
        marker=dict(color=SEQ[3], cornerradius=4),
        hovertemplate='%{y}: %{x:,.0f}<extra></extra>',
    ))
    _base_layout(fig, height=440)
    fig.update_layout(hovermode='closest',
                      margin=dict(l=180, r=24, t=28, b=40))
    return _bar(fig)


def _chart_cv_results(df_cv):
    df = df_cv.sort_values('Fold')
    colors = [STATUS['warning'] if h > 1 else SEQ[3] for h in df['Holiday_Weeks']]
    fig = go.Figure(go.Bar(
        x=[f"Fold {int(f)}<br><span style='font-size:10px'>{s}</span>"
           for f, s in zip(df['Fold'], df['Test_Start'])],
        y=df['WMAE'],
        marker=dict(color=colors, cornerradius=4),
        text=[f'{v:,.0f}' for v in df['WMAE']], textposition='outside',
        textfont=dict(color=INK_2, size=11),
        customdata=df[['MAPE', 'Holiday_Weeks']].values,
        hovertemplate='WMAE: %{y:,.0f}<br>MAPE: %{customdata[0]:.1f}% · '
                      '%{customdata[1]} holiday weeks<extra></extra>',
        width=0.55,
    ))
    mean_w = df['WMAE'].mean()
    fig.add_hline(y=mean_w, line_width=1.5, line_dash='dash', line_color=BASELINE)
    _base_layout(fig)
    fig.update_layout(hovermode='closest')
    return _bar(fig)


def _chart_run_history(df_runs):
    df = df_runs.reset_index(drop=True)
    fig = go.Figure(go.Scatter(
        x=list(range(1, len(df) + 1)), y=df['WMAE'],
        mode='lines+markers',
        line=dict(color=SERIES_1, width=2), marker=dict(size=8),
        customdata=df[['timestamp', 'model', 'run_mode']].astype(str).values,
        hovertemplate='Run %{x}: WMAE %{y:,.0f}<br>%{customdata[0]}<br>'
                      '%{customdata[1]} (%{customdata[2]})<extra></extra>',
    ))
    _base_layout(fig)
    fig.update_layout(hovermode='closest',
                      xaxis=dict(title=dict(text='Run #', font=dict(size=11, color=MUTED)),
                                 dtick=1))
    return fig


# ==========================================================================
# HTML PIECES
# ==========================================================================

def _restock_table(detailed, n=10):
    if detailed is None or len(detailed) == 0:
        return '<p class="empty">No items need restocking in the evaluation window.</p>'

    rows = []
    for _, r in detailed.head(n).iterrows():
        sr = r['Stockout_Rate']
        if sr >= 50:
            color, lbl = STATUS['critical'], 'Critical'
        elif sr >= 30:
            color, lbl = STATUS['serious'], 'High'
        else:
            color, lbl = STATUS['warning'], 'Watch'
        rows.append(f'''<tr>
          <td>Store {int(r['Store'])}</td><td>Dept {int(r['Dept'])}</td>
          <td>{r['ABC']}{r['XYZ']}</td>
          <td class="num">{_fmt_money(r['Avg_Weekly_Sales'])}</td>
          <td class="num">{sr:.0f}%</td>
          <td class="num">{r['Service_Level']:.0f}%</td>
          <td class="num">{_fmt_money(r['Restock_Quantity'])}</td>
          <td><span class="chip" style="--c:{color}">{lbl}</span></td>
        </tr>''')

    return f'''<table class="tbl">
      <thead><tr>
        <th>Store</th><th>Department</th><th>Class</th>
        <th class="num">Avg weekly sales</th>
        <th class="num">Stockout rate</th><th class="num">Service level</th>
        <th class="num">Restock value</th><th>Status</th>
      </tr></thead>
      <tbody>{''.join(rows)}</tbody>
    </table>'''


def _build_insights(df, metrics, inventory):
    insights = []
    y, p = df['Actual'], df['Predicted']
    bias = (p.sum() / y.sum() - 1) * 100
    wape = (y - p).abs().sum() / y.abs().sum() * 100

    insights.append(
        f'Network-wide forecast accuracy is <b>{100 - wape:.1f}%</b> (WAPE {wape:.1f}%), '
        f'with total forecast off by <b>{bias:+.1f}%</b> — '
        f'{"within the ±2% safe band for inventory planning" if abs(bias) <= 2 else "bias correction recommended before planning use"}.')

    dept_err = (df.assign(e=(y - p).abs()).groupby('Dept')['e'].sum().sort_values(ascending=False))
    top3_share = dept_err.head(3).sum() / dept_err.sum() * 100
    top3 = ', '.join(f'Dept {int(d)}' for d in dept_err.head(3).index)
    insights.append(
        f'Error is concentrated: 3 departments (<b>{top3}</b>) account for <b>{top3_share:.0f}%</b> '
        f'of total error across {df["Dept"].nunique()} departments — targeted improvements there pay off most.')

    if df['IsHoliday'].any():
        mae_h = (y - p).abs()[df['IsHoliday']].mean()
        mae_n = (y - p).abs()[~df['IsHoliday']].mean()
        insights.append(
            f'Holiday weeks run at MAE <b>{_fmt_money(mae_h)}</b> vs {_fmt_money(mae_n)} for regular weeks '
            f'({mae_h / mae_n:.1f}x) — note the test window only contains Labor Day; '
            f'Thanksgiving/Christmas performance is assessed via cross-validation instead.')

    detailed = inventory.get('store_dept_detailed') if inventory else None
    if detailed is not None and len(detailed):
        a_items = detailed[detailed['ABC'] == 'A']
        if len(a_items):
            r = a_items.iloc[0]
            insights.append(
                f'<b>{len(detailed)}</b> Store-Dept series need restocking; most notable is '
                f'<b>Store {int(r["Store"])}/Dept {int(r["Dept"])}</b> — an A-class series '
                f'({_fmt_money(r["Avg_Weekly_Sales"])}/week) stocking out '
                f'{r["Stockout_Rate"]:.0f}% of weeks — direct revenue risk.')
        else:
            insights.append(f'<b>{len(detailed)}</b> Store-Dept series need restocking, '
                            f'all in low-value B/C classes.')

    return insights


def _page(title, eyebrow, meta, nav_html, body, footer):
    return f'''<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{title}</title>
<style>{_page_css()}</style></head>
<body><div class="wrap">
  <header>
    <div>
      <div class="eyebrow">{eyebrow}</div>
      <h1>{title}</h1>
      <div class="meta">{meta}</div>
    </div>
    <nav>{nav_html}</nav>
  </header>
  {body}
  <footer>{footer}</footer>
</div></body></html>'''


# ==========================================================================
# PUBLIC BUILDERS
# ==========================================================================

def build_dashboard(eval_data, metrics, best_model_name, feature_importance_df=None,
                    inventory=None, comparison_results=None,
                    out_path='results/dashboard.html'):
    """BUSINESS ANALYTICS page."""
    if go is None:
        logger.warning('Plotly not installed - skipping dashboard.')
        return None

    logger.info('Building analyst dashboard...')
    df = _eval_df(eval_data)
    y, p = df['Actual'], df['Predicted']
    period = f"{df['Date'].min():%b %d, %Y} – {df['Date'].max():%b %d, %Y}"
    n_weeks = df['Date'].nunique()
    bias = (p.sum() / y.sum() - 1) * 100
    wape = (y - p).abs().sum() / y.abs().sum() * 100

    detailed = inventory.get('store_dept_detailed') if inventory else None
    full = inventory.get('store_dept_full') if inventory else None
    n_restock = len(detailed) if detailed is not None else 0
    service = (100 - full['Stockout_Rate'].mean()) if full is not None and len(full) else None

    kpis = [
        _kpi_card('Total sales (test window)', _fmt_money(y.sum() / 1e6, 1) + 'M',
                  f'{n_weeks} weeks · {df["Store"].nunique()} stores'),
        _kpi_card('Forecast accuracy', f'{100 - wape:.1f}%', '1 − Σ|error| / Σsales',
                  tone='good' if wape < 10 else 'warning'),
        _kpi_card('Total forecast bias', f'{bias:+.1f}%', 'negative = under-forecasting',
                  tone='good' if abs(bias) <= 2 else 'warning'),
        _kpi_card('WMAE', _fmt_money(metrics['WMAE']), 'primary metric · holidays ×5'),
    ]
    if service is not None:
        kpis.append(_kpi_card('Avg service level', f'{service:.1f}%', 'ABC-XYZ inventory policy',
                              tone='good' if service >= 95 else 'serious'))
    kpis.append(_kpi_card('Series needing restock', f'{n_restock:,}',
                          'of ' + (f'{len(full):,}' if full is not None else '—') + ' Store-Dept series',
                          tone='warning' if n_restock else 'good'))

    insights = _build_insights(df, metrics, inventory)
    insights_html = '<div class="insights"><h2>💡 Key takeaways</h2><ul>' + \
                    ''.join(f'<li>{s}</li>' for s in insights) + '</ul></div>'

    cards = [
        _card('Weekly sales: Actual vs Forecast',
              'Network total · dotted verticals mark holiday weeks · filter by store cluster (top right)',
              _chart_forecast_timeseries(df), size='full', include_js=True),
        _card('Forecast error (MAE) by store cluster',
              'Clusters from K-Means on size, type and volatility',
              _chart_error_by_cluster(df)),
        _card('Total sales per store: Forecast vs Actual',
              'Each dot is a store · below the diagonal = under-forecast',
              _chart_store_scatter(df)),
    ]
    if full is not None and len(full):
        cards.append(_card('ABC × XYZ inventory matrix',
                           'Store-Dept pairs and revenue share per cell - drives safety-stock levels',
                           _chart_abc_xyz(full)))
    cards.append(_card('Error concentration by department',
                       'Share of network-wide absolute error - where model improvements pay off most',
                       _chart_error_concentration(df)))
    if comparison_results:
        cards.append(_card('Model comparison by WMAE', 'Lower is better · dark bar = selected model',
                           _chart_model_comparison(pd.DataFrame(comparison_results)), size='full'))

    table_card = f'''<div class="card full">
      <h2 class="plain">Top restock priorities</h2>
      <div class="pad">{_restock_table(detailed)}</div>
      <div class="note">Series with ≥4 weeks of data and positive sales only. Values are in dollars
        (the dataset has no unit prices). Full list: sheet "Restock_Priority" in forecast_report.xlsx</div>
    </div>'''

    body = f'''
  <div class="section-label">Key metrics</div>
  <div class="kpis">{''.join(kpis)}</div>
  <div class="section-label">Insights</div>
  {insights_html}
  {_section('Forecast performance', ''.join(cards))}
  {_section('Inventory actions', table_card)}'''

    footer = (f'Model: {best_model_name} · Test window: {period} · '
              f'WMAE = weighted MAE (holiday weeks ×5, Kaggle Walmart standard) · '
              f'Detailed data: results/reports/forecast_report.xlsx')

    page = _page(
        title='Walmart Demand Forecasting — Analytics Report',
        eyebrow='Business Analytics',
        meta=f'Model: <b>{best_model_name}</b> · Test window: {period} · Updated: {datetime.now():%b %d, %Y %H:%M}',
        nav_html='<a href="model_report.html">Model diagnostics →</a>',
        body=body, footer=footer)

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(page, encoding='utf-8')
    logger.info(f'  ✓ Analyst dashboard saved: {out}')
    return str(out)


def _extract_model_info(model_instance, model_name):
    info = {}
    m = getattr(model_instance, 'model', None)
    if m is None:
        return info
    try:
        params = m.get_params()
        for k in ['n_estimators', 'iterations', 'learning_rate', 'num_leaves',
                  'max_depth', 'depth', 'objective', 'loss_function',
                  'subsample', 'colsample_bytree']:
            if k in params and params[k] is not None:
                info[k] = params[k]
    except Exception:
        pass
    for attr in ['best_iteration_', 'best_iteration']:
        v = getattr(m, attr, None)
        if v:
            info['best_iteration (early stopping)'] = v
            break
    n_feat = len(getattr(model_instance, 'feature_names_', []) or [])
    if n_feat:
        info['n_features'] = n_feat
    return info


def build_model_report(eval_data, metrics, best_model_name, model_instance=None,
                       split_info=None, feature_importance_df=None,
                       comparison_results=None,
                       out_path='results/model_report.html'):
    """MODEL DIAGNOSTICS page for data scientists."""
    if go is None:
        logger.warning('Plotly not installed - skipping model report.')
        return None

    logger.info('Building model diagnostics report...')
    df = _eval_df(eval_data)
    y, p = df['Actual'], df['Predicted']
    resid = y - p
    period = f"{df['Date'].min():%b %d, %Y} – {df['Date'].max():%b %d, %Y}"
    bias = (p.sum() / y.sum() - 1) * 100

    kpis = [
        _kpi_card('WMAE', _fmt_money(metrics['WMAE']), 'primary metric · holidays ×5'),
        _kpi_card('MAE', _fmt_money(metrics['MAE'])),
        _kpi_card('RMSE', _fmt_money(metrics['RMSE']), 'outlier-sensitive'),
        _kpi_card('MAPE', _fmt_pct(metrics.get('MAPE')), 'sales > $100 only'),
        _kpi_card('sMAPE', _fmt_pct(metrics.get('sMAPE'))),
        _kpi_card('Total bias', f'{bias:+.2f}%',
                  f'median residual {_fmt_money(float(resid.median()))}'),
    ]

    model_info = _extract_model_info(model_instance, best_model_name) if model_instance else {}
    info_rows = ''.join(f'<tr><td>{k}</td><td class="num">{v}</td></tr>' for k, v in model_info.items())
    split_rows = ''
    if split_info:
        for name, (d0, d1, n) in split_info.items():
            split_rows += (f'<tr><td>{name}</td>'
                           f'<td class="num">{pd.Timestamp(d0):%Y-%m-%d} – {pd.Timestamp(d1):%Y-%m-%d}</td>'
                           f'<td class="num">{n:,}</td></tr>')

    config_card = f'''<div class="card">
      <h2 class="plain">Model configuration — {best_model_name}</h2>
      <div class="pad"><table class="tbl">
        <thead><tr><th>Parameter</th><th class="num">Value</th></tr></thead>
        <tbody>{info_rows or '<tr><td colspan="2" class="empty">No information</td></tr>'}</tbody>
      </table></div>
      <div class="note">Trained with sample_weight = 5 on holiday weeks (matches WMAE) ·
        early stopping on validation · feature statistics fitted on train only (leak-free)</div>
    </div>'''
    split_card = f'''<div class="card">
      <h2 class="plain">Data split (time-based, 70/15/15)</h2>
      <div class="pad"><table class="tbl">
        <thead><tr><th>Split</th><th class="num">Date range</th><th class="num">Rows</th></tr></thead>
        <tbody>{split_rows or '<tr><td colspan="3" class="empty">No information</td></tr>'}</tbody>
      </table></div>
      <div class="note">No date overlap (asserted in the pipeline) · features are lag-based,
        so no future information reaches the model</div>
    </div>'''

    # ---- Section 1: error diagnostics ----
    resid_pct_out = ((y - p).abs() > 5000).mean() * 100
    diag_cards = [
        _card('Error (MAE) by test week',
              'Dotted verticals mark holiday weeks - spot where the model struggles',
              _chart_mae_by_week(df), size='full', include_js=True),
        _card('Residual distribution (Actual − Forecast)',
              f'Clipped at ±$5,000 for display ({resid_pct_out:.1f}% outside) · '
              f'dashed line = median · right skew = under-forecasting',
              _chart_residual_hist(df)),
        _card('Residual vs forecast level',
              f'Random sample of {min(8000, len(df)):,} points · funnel shape = variance grows with scale',
              _chart_residual_vs_predicted(df)),
        _card('MAE: holiday vs regular weeks',
              'WMAE weighs holiday errors 5× - the key segment to watch',
              _chart_holiday_error(df)),
        _card('Top 10 departments by MAE',
              'Departments with ≥30 test observations only',
              _chart_worst_depts(df)),
    ]

    # ---- Section 2: validation & experiments ----
    val_cards = []
    cv_path = Path('results/cv_results.csv')
    if cv_path.exists():
        try:
            df_cv = pd.read_csv(cv_path)
            if len(df_cv):
                val_cards.append(_card(
                    'Rolling-origin cross-validation (WMAE per fold)',
                    'Orange = fold with major holidays (Thanksgiving/Christmas) in its test window · dashed line = mean',
                    _chart_cv_results(df_cv)))
        except Exception:
            pass
    df_runs = load_runs()
    if len(df_runs) >= 2:
        val_cards.append(_card('WMAE across pipeline runs',
                               'Lightweight experiment tracking from results/runs.jsonl',
                               _chart_run_history(df_runs)))

    # ---- Section 3: model comparison ----
    df_cmp = None
    if comparison_results:
        df_cmp = pd.DataFrame(comparison_results)
    else:
        cmp_path = Path('results/model_comparison.csv')
        if cmp_path.exists():
            try:
                df_cmp = pd.read_csv(cmp_path)
            except Exception:
                pass
    cmp_cards = []
    if df_cmp is not None and 'WMAE' in df_cmp.columns:
        cmp_cards.append(_card('Model comparison by WMAE',
                               'Lower is better · dark bar = selected model',
                               _chart_model_comparison(df_cmp)))
        tt = _chart_train_time(df_cmp)
        if tt is not None:
            cmp_cards.append(_card('Training time', 'Accuracy vs compute-cost trade-off', tt))

    # ---- Section 4: feature importance ----
    fi_cards = []
    if feature_importance_df is not None and len(feature_importance_df):
        fi_cards.append(_card(f'Top 15 feature importances ({best_model_name})',
                              'Contribution of each feature to tree splits',
                              _chart_feature_importance(feature_importance_df, best_model_name),
                              size='full'))

    sections = [
        f'''<div class="section-label">Test-set metrics</div>
  <div class="kpis">{''.join(kpis)}</div>''',
        _section('Experiment setup', config_card + split_card),
        _section('Error diagnostics', ''.join(diag_cards)),
    ]
    if val_cards:
        sections.append(_section('Validation & experiments', ''.join(val_cards)))
    if cmp_cards:
        sections.append(_section('Model comparison', ''.join(cmp_cards)))
    if fi_cards:
        sections.append(_section('Feature importance', ''.join(fi_cards)))

    body = '\n'.join(sections)

    footer = (f'Test window: {period} · {len(df):,} observations · 1-step-ahead forecast '
              f'(uses previous week\'s lag) · Model comparison from the latest compare run · '
              f'CV from validate.py')

    page = _page(
        title='Model Diagnostics — Data Science Report',
        eyebrow='Data Science',
        meta=f'Model: <b>{best_model_name}</b> · Test window: {period} · Updated: {datetime.now():%b %d, %Y %H:%M}',
        nav_html='<a href="dashboard.html">← Analytics report</a>',
        body=body, footer=footer)

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(page, encoding='utf-8')
    logger.info(f'  ✓ Model report saved: {out}')
    return str(out)


def build_excel_report(metrics, best_model_name, eval_data, inventory=None,
                       comparison_results=None, feature_importance_df=None,
                       out_path='results/reports/forecast_report.xlsx'):
    """Formatted multi-sheet Excel workbook."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("openpyxl not installed - skipping Excel report.")
        return None

    logger.info('Building Excel report...')

    HEADER_FILL = PatternFill('solid', fgColor='1C5CAB')
    HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
    TITLE_FONT = Font(bold=True, size=14, color='0B0B0B')
    THIN = Border(bottom=Side(style='thin', color='E1E0D9'))

    def style_table(ws, df, start_row=1):
        for j, col in enumerate(df.columns, start=1):
            c = ws.cell(row=start_row, column=j)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal='center', vertical='center')
        for j, col in enumerate(df.columns, start=1):
            width = max(len(str(col)) + 2, 12)
            ws.column_dimensions[get_column_letter(j)].width = min(width, 28)
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
        if len(df) > 0:
            last_col = get_column_letter(len(df.columns))
            ws.auto_filter.ref = f'A{start_row}:{last_col}{start_row + len(df)}'
        for row in ws.iter_rows(min_row=start_row + 1, max_row=start_row + len(df)):
            for c in row:
                c.border = THIN
                if isinstance(c.value, float):
                    c.number_format = '#,##0.0'

    def add_colorscale(ws, df, col_name, start_row=1, reverse=False):
        if col_name not in df.columns or len(df) == 0:
            return
        j = list(df.columns).index(col_name) + 1
        letter = get_column_letter(j)
        ref = f'{letter}{start_row + 1}:{letter}{start_row + len(df)}'
        lo, hi = ('D03B3B', '0CA30C') if reverse else ('CDE2FB', '1C5CAB')
        ws.conditional_formatting.add(ref, ColorScaleRule(
            start_type='min', start_color=lo, end_type='max', end_color=hi))

    y_true = np.asarray(eval_data['y_true'])
    y_pred = np.asarray(eval_data['y_pred'])
    dates = pd.to_datetime(pd.Series(eval_data['Date']))

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    try:
        with pd.ExcelWriter(out, engine='openpyxl') as writer:
            # ---------- Sheet 1: Overview ----------
            summary_rows = [
                ('Best model', best_model_name),
                ('Test window', f"{dates.min():%Y-%m-%d} – {dates.max():%Y-%m-%d}"),
                ('Test observations', f'{len(y_true):,}'),
                ('WMAE (holidays x5) - primary metric', round(float(metrics['WMAE']), 1)),
                ('MAE', round(float(metrics['MAE']), 1)),
                ('RMSE', round(float(metrics['RMSE']), 1)),
                ('MAPE (%)', round(float(metrics['MAPE']), 2)),
                ('Total actual sales ($)', round(float(y_true.sum()), 0)),
                ('Total forecast sales ($)', round(float(y_pred.sum()), 0)),
                ('Total forecast bias (%)', round(float((y_pred.sum() / y_true.sum() - 1) * 100), 2)),
                ('Report generated', datetime.now().strftime('%Y-%m-%d %H:%M')),
            ]
            df_sum = pd.DataFrame(summary_rows, columns=['Metric', 'Value'])
            df_sum.to_excel(writer, sheet_name='Overview', index=False, startrow=2)
            ws = writer.sheets['Overview']
            ws['A1'] = 'WALMART DEMAND FORECASTING REPORT'
            ws['A1'].font = TITLE_FONT
            style_table(ws, df_sum, start_row=3)
            ws.column_dimensions['A'].width = 38
            ws.column_dimensions['B'].width = 26

            # ---------- Sheet 2: Model comparison ----------
            if comparison_results:
                df_cmp = pd.DataFrame(comparison_results)
                cols = [c for c in ['Model', 'WMAE', 'MAE', 'RMSE', 'MAPE', 'sMAPE'] if c in df_cmp.columns]
                df_cmp = df_cmp[cols].sort_values('WMAE').reset_index(drop=True)
                df_cmp.to_excel(writer, sheet_name='Model_Comparison', index=False)
                ws = writer.sheets['Model_Comparison']
                style_table(ws, df_cmp)
                add_colorscale(ws, df_cmp, 'WMAE')

            # ---------- Sheet 3: Store inventory ----------
            if inventory and inventory.get('store_summary') is not None:
                df_store = inventory['store_summary'].copy()
                df_store.to_excel(writer, sheet_name='Store_Inventory', index=False)
                ws = writer.sheets['Store_Inventory']
                style_table(ws, df_store)
                add_colorscale(ws, df_store, 'Avg_Error_Pct')

            # ---------- Sheet 4: Restock priority ----------
            if inventory and inventory.get('store_dept_detailed') is not None:
                df_re = inventory['store_dept_detailed'].copy()
                df_re.to_excel(writer, sheet_name='Restock_Priority', index=False)
                ws = writer.sheets['Restock_Priority']
                style_table(ws, df_re)
                add_colorscale(ws, df_re, 'Stockout_Rate')
                add_colorscale(ws, df_re, 'Priority_Score')

            # ---------- Sheet 5: All Store-Dept (ABC-XYZ) ----------
            if inventory and inventory.get('store_dept_full') is not None:
                df_full = inventory['store_dept_full'].copy()
                df_full.to_excel(writer, sheet_name='ABC_XYZ_All', index=False)
                ws = writer.sheets['ABC_XYZ_All']
                style_table(ws, df_full)
                add_colorscale(ws, df_full, 'Service_Level', reverse=True)

            # ---------- Sheet 6: Feature importance ----------
            if feature_importance_df is not None and len(feature_importance_df):
                df_imp = feature_importance_df.copy()
                df_imp.to_excel(writer, sheet_name='Feature_Importance', index=False)
                ws = writer.sheets['Feature_Importance']
                style_table(ws, df_imp)
                add_colorscale(ws, df_imp, 'Importance')

        logger.info(f'  ✓ Excel report saved: {out}')
        return str(out)

    except PermissionError:
        logger.warning(f'  {out} is open in Excel - close it and rerun.')
        return None
